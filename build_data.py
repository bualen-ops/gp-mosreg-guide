#!/usr/bin/env python3
import json
from pathlib import Path

import openpyxl


SOURCE_XLSX = Path("/Users/alenbubnovskiy/Downloads/Реестр_Вводных_2026.xlsx")
OUTPUT_JSON = Path(__file__).resolve().parent / "data.json"

FIELD_MAP = {
    "Направление": "direction",
    "Код мероприятия": "eventCode",
    "Субсидия": "subsidy",
    "Фед. субсидия": "federalSubsidy",
    "ОМСУ": "region",
    "Наименование объекта": "objectName",
    "Вид работ": "workType",
    "Тип объекта": "objectType",
    "Вводные": "intro",
    "Срок реализации": "timeline",
    "Год ввода": "commissioningYear",
    "Охват": "coverage",
    "Мощность, куб.м./сут": "capacityPerDay",
    "Протяженность, метр": "lengthMeters",
    "№ контракта": "contractNumber",
    "Подрядная \nорганизация": "contractor",
    "РП": "manager",
    "Комментарий": "comment",
}


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        compact = " ".join(value.split())
        return compact or None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def main():
    if not SOURCE_XLSX.exists():
        raise FileNotFoundError(f"Excel file not found: {SOURCE_XLSX}")

    workbook = openpyxl.load_workbook(SOURCE_XLSX, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_index = {name: idx for idx, name in enumerate(header_row) if name}

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        object_name = normalize_value(row[header_index["Наименование объекта"]])
        region = normalize_value(row[header_index["ОМСУ"]])
        if not object_name:
            continue

        record = {"objectName": object_name, "region": region}
        for source_field, target_field in FIELD_MAP.items():
            idx = header_index.get(source_field)
            if idx is None:
                continue
            record[target_field] = normalize_value(row[idx])

        searchable_parts = [
            record.get("region"),
            record.get("objectName"),
            record.get("workType"),
            record.get("objectType"),
            record.get("eventCode"),
        ]
        record["searchText"] = " ".join([p.lower() for p in searchable_parts if p])
        records.append(record)

    payload = {
        "source": str(SOURCE_XLSX),
        "totalRecords": len(records),
        "regions": sorted({r["region"] for r in records if r.get("region")}),
        "items": records,
    }

    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Saved {len(records)} records to {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
